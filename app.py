"""
Kibar Holding – İSGÇ OKR Takip Sistemi
Flask backend: auth, veri giriş, dashboard API, Excel export
"""

from flask import Flask, request, jsonify, send_file, render_template, session
import sqlite3, hashlib, hmac, secrets, json, io, os
from datetime import datetime
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__, template_folder=os.path.join(os.path.dirname(__file__), 'templates'))
app.secret_key = secrets.token_hex(32)

DB = "isgc.db"

# ─── OKR YAPISI ──────────────────────────────────────────────────────────────

OKR_STRUKTUR = [
    {"no": 1, "okr": "LİDERLİK, YÖNETİM, SORUMLULUK", "krs": [
        {"no": 1, "kr": "Yöneticilerin Planlı İSGÇ Aktivitelerine Katılım Oranı", "birim": "%"},
        {"no": 2, "kr": "Yöneticilerin İSGÇ Kurul Toplantılarına Katılım Oranı", "birim": "%"},
        {"no": 3, "kr": "İSGÇ Liderler Saha Turu Katılım Oranı", "birim": "%"},
        {"no": 4, "kr": "Yıllık Bölüm İSGÇ Hedeflerinin Belirlenmesi Ve İlgili Yöneticilerin Performans Hedefi Olarak Verilme Oranı", "birim": "%"},
        {"no": 5, "kr": "Üst Yönetim Tarafından Şirket İSGÇ Hedeflerinin Aylık Bazda Planlı Olarak Gözden Geçirilme Oranı", "birim": "%"},
    ]},
    {"no": 2, "okr": "YASALARA ve STANDARTLARA UYUM", "krs": [
        {"no": 1, "kr": "İş Sağlığı Mevzuatı Uyum Oranı", "birim": "%"},
        {"no": 2, "kr": "İş Güvenliği Mevzuatı Uyum Oranı", "birim": "%"},
        {"no": 3, "kr": "Çevre Mevzuatı Uyum Oranı", "birim": "%"},
        {"no": 4, "kr": "İş Sağlığı Mevzuatı Uyumu Geciken Aksiyon Sayısı", "birim": "sayı"},
        {"no": 5, "kr": "İş Güvenliği Mevzuatı Uyumu Geciken Aksiyon Sayısı", "birim": "sayı"},
        {"no": 6, "kr": "Çevre Mevzuatı Uyumu Geciken Aksiyon Sayısı", "birim": "sayı"},
    ]},
    {"no": 3, "okr": "RİSK YÖNETİMİ", "krs": [
        {"no": 1, "kr": "Operasyon Bazlı İş Sağlığı Risk Değerlendirme Tamamlanma Oranı", "birim": "%"},
        {"no": 2, "kr": "Çok Yüksek İş Sağlığı Risk Sayısı", "birim": "sayı"},
        {"no": 3, "kr": "Yüksek İş Sağlığı Risk Sayısı", "birim": "sayı"},
    ]},
    {"no": 4, "okr": "EĞİTİM, ÖĞRETİM VE FARKINDALIK", "krs": [
        {"no": 1, "kr": "Çalışan Başına Yıllık İSGÇ Eğitim Saati", "birim": "saat"},
        {"no": 2, "kr": "İSGÇ Eğitim Planı Tamamlanma Oranı", "birim": "%"},
    ]},
    {"no": 5, "okr": "OPERASYONEL İSGÇ STANDARTLARI", "krs": [
        {"no": 1, "kr": "Kayıp Günlü Kaza Sıklık Oranı (Ş&A)", "birim": "oran"},
        {"no": 2, "kr": "Kaza Ağırlık Oranı (Ş&A)", "birim": "oran"},
        {"no": 3, "kr": "Ramak Kala Sayısı (Ş&A)", "birim": "sayı"},
    ]},
    {"no": 6, "okr": "DEĞİŞİM YÖNETİMİ", "krs": [
        {"no": 1, "kr": "Değişim Yönetimi Prosedürü Uygulanma Oranı", "birim": "%"},
    ]},
    {"no": 7, "okr": "HİZMET VE ÜRÜN ALIMI", "krs": [
        {"no": 1, "kr": "Yüklenici İSGÇ Denetim Tamamlanma Oranı", "birim": "%"},
    ]},
    {"no": 8, "okr": "ACİL DURUM YÖNETİMİ", "krs": [
        {"no": 1, "kr": "Acil Durum Tatbikat Planı Tamamlanma Oranı", "birim": "%"},
    ]},
    {"no": 9, "okr": "KAZA, OLAY ARAŞTIRMA", "krs": [
        {"no": 1, "kr": "Kaza/Olay Araştırma Tamamlanma Oranı", "birim": "%"},
        {"no": 2, "kr": "Kaza/Olay Aksiyonlarının Zamanında Kapanma Oranı", "birim": "%"},
    ]},
    {"no": 10, "okr": "YENİ PROJELER VE TASFİYELER", "krs": [
        {"no": 1, "kr": "Yeni Proje İSGÇ Değerlendirme Tamamlanma Oranı", "birim": "%"},
    ]},
    {"no": 11, "okr": "GÖZLEM, SÜREKLİ İYİLEŞTİRME, İLETİŞİM", "krs": [
        {"no": 1, "kr": "İSGÇ Gözlem Planı Tamamlanma Oranı", "birim": "%"},
        {"no": 2, "kr": "Kapatılan İyileştirme Aksiyonu Oranı", "birim": "%"},
    ]},
]

COMPANIES = [
    "ASSAN ALÜMİNYUM",
    "ASSAN HANİL",
    "ASSAN LİMAN",
    "ASSAN LOJİSTİK",
    "ASSAN PANEL",
    "İSPAK ESNEK AMBALAJ",
]

MONTHS = ["JAN","FEB","MAR","APR","MAY","JUN","JUL","AGU","SEP","OKT","NOV","DEC"]
MONTH_TR = ["Ocak","Şubat","Mart","Nisan","Mayıs","Haziran","Temmuz","Ağustos","Eylül","Ekim","Kasım","Aralık"]

# ─── VERİTABANI ──────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()

    c.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'company',
        company TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS entries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company TEXT NOT NULL,
        yil INTEGER NOT NULL,
        ay TEXT NOT NULL,
        okr_no INTEGER NOT NULL,
        kr_no INTEGER NOT NULL,
        deger REAL,
        giren_user TEXT,
        created_at TEXT DEFAULT (datetime('now')),
        updated_at TEXT DEFAULT (datetime('now')),
        UNIQUE(company, yil, ay, okr_no, kr_no)
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS sessions (
        token TEXT PRIMARY KEY,
        user_id INTEGER,
        expires_at TEXT
    )""")

    conn.commit()

    # Varsayılan kullanıcılar
    default_users = [
        ("admin",         "Admin2026!",  "admin",   None),
        ("assan_al",      "AlPass26!",   "company", "ASSAN ALÜMİNYUM"),
        ("assan_hanil",   "HanPass26!",  "company", "ASSAN HANİL"),
        ("assan_liman",   "LimPass26!",  "company", "ASSAN LİMAN"),
        ("assan_loj",     "LojPass26!",  "company", "ASSAN LOJİSTİK"),
        ("assan_panel",   "PanPass26!",  "company", "ASSAN PANEL"),
        ("ispak",         "IspPass26!",  "company", "İSPAK ESNEK AMBALAJ"),
    ]
    for username, password, role, company in default_users:
        try:
            ph = hash_password(password)
            c.execute("INSERT INTO users (username,password_hash,role,company) VALUES (?,?,?,?)",
                      (username, ph, role, company))
        except:
            pass
    conn.commit()
    conn.close()

def hash_password(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def verify_password(pw, hashed):
    return hmac.compare_digest(hash_password(pw), hashed)

# ─── AUTH ─────────────────────────────────────────────────────────────────────

def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get("Authorization","").replace("Bearer ","")
        if not token:
            token = request.cookies.get("token","")
        if not token:
            return jsonify({"error": "Yetkisiz erişim"}), 401
        conn = get_db()
        row = conn.execute(
            "SELECT u.* FROM sessions s JOIN users u ON s.user_id=u.id WHERE s.token=? AND s.expires_at > datetime('now')",
            (token,)
        ).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Oturum süresi doldu"}), 401
        request.user = dict(row)
        return f(*args, **kwargs)
    return decorated

def require_admin(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get("Authorization","").replace("Bearer ","")
        if not token:
            token = request.cookies.get("token","")
        conn = get_db()
        row = conn.execute(
            "SELECT u.* FROM sessions s JOIN users u ON s.user_id=u.id WHERE s.token=? AND s.expires_at > datetime('now')",
            (token,)
        ).fetchone()
        conn.close()
        if not row or dict(row)["role"] != "admin":
            return jsonify({"error": "Yönetici yetkisi gerekli"}), 403
        request.user = dict(row)
        return f(*args, **kwargs)
    return decorated

# ─── SAYFALAR ─────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/dashboard")
def dashboard():
    return render_template("dashboard.html")

@app.route("/giris")
def giris_page():
    return render_template("giris.html")

# ─── API: AUTH ────────────────────────────────────────────────────────────────

@app.route("/api/login", methods=["POST"])
def login():
    data = request.json or {}
    username = data.get("username","").strip()
    password = data.get("password","")

    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    if not user or not verify_password(password, user["password_hash"]):
        conn.close()
        return jsonify({"error": "Kullanıcı adı veya şifre hatalı"}), 401

    token = secrets.token_hex(32)
    conn.execute(
        "INSERT INTO sessions(token,user_id,expires_at) VALUES(?,?,datetime('now','+8 hours'))",
        (token, user["id"])
    )
    conn.commit()
    conn.close()

    return jsonify({
        "token": token,
        "role": user["role"],
        "company": user["company"],
        "username": user["username"]
    })

@app.route("/api/logout", methods=["POST"])
@require_auth
def logout():
    token = request.headers.get("Authorization","").replace("Bearer ","") or request.cookies.get("token","")
    conn = get_db()
    conn.execute("DELETE FROM sessions WHERE token=?", (token,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/me")
@require_auth
def me():
    return jsonify({
        "username": request.user["username"],
        "role": request.user["role"],
        "company": request.user["company"]
    })

# ─── API: VERİ GİRİŞİ ────────────────────────────────────────────────────────

@app.route("/api/okr-struktur")
@require_auth
def okr_struktur():
    return jsonify(OKR_STRUKTUR)

@app.route("/api/entries", methods=["POST"])
@require_auth
def save_entries():
    """Aylık toplu veri kaydetme"""
    data = request.json or {}
    company = request.user["company"]
    if request.user["role"] == "admin":
        company = data.get("company", company)

    yil  = int(data.get("yil", datetime.now().year))
    ay   = data.get("ay", "")
    rows = data.get("rows", [])  # [{okr_no, kr_no, deger}]

    if not ay or ay not in MONTHS:
        return jsonify({"error": "Geçersiz ay"}), 400
    if not company:
        return jsonify({"error": "Şirket belirtilmedi"}), 400

    conn = get_db()
    saved = 0
    for row in rows:
        try:
            conn.execute("""
                INSERT INTO entries (company,yil,ay,okr_no,kr_no,deger,giren_user,updated_at)
                VALUES (?,?,?,?,?,?,?,datetime('now'))
                ON CONFLICT(company,yil,ay,okr_no,kr_no)
                DO UPDATE SET deger=excluded.deger, updated_at=excluded.updated_at, giren_user=excluded.giren_user
            """, (company, yil, ay, row["okr_no"], row["kr_no"], row.get("deger"), request.user["username"]))
            saved += 1
        except Exception as e:
            pass
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "saved": saved})

@app.route("/api/entries")
@require_auth
def get_entries():
    company = request.args.get("company")
    yil  = request.args.get("yil", datetime.now().year)
    ay   = request.args.get("ay")

    if request.user["role"] == "company":
        company = request.user["company"]

    conn = get_db()
    q = "SELECT * FROM entries WHERE yil=?"
    params = [yil]
    if company:
        q += " AND company=?"; params.append(company)
    if ay:
        q += " AND ay=?"; params.append(ay)
    rows = [dict(r) for r in conn.execute(q, params).fetchall()]
    conn.close()
    return jsonify(rows)

# ─── API: DASHBOARD ───────────────────────────────────────────────────────────

@app.route("/api/dashboard/summary")
@require_auth
def dashboard_summary():
    yil = request.args.get("yil", datetime.now().year)
    conn = get_db()

    result = []
    for company in COMPANIES:
        rows = conn.execute(
            "SELECT ay, AVG(deger) as avg_deger FROM entries WHERE company=? AND yil=? AND deger IS NOT NULL GROUP BY ay",
            (company, yil)
        ).fetchall()

        monthly = {r["ay"]: round(r["avg_deger"], 2) for r in rows}
        ytd_vals = [v for v in monthly.values()]
        ytd = round(sum(ytd_vals)/len(ytd_vals), 2) if ytd_vals else None

        result.append({
            "company": company,
            "monthly": monthly,
            "ytd": ytd,
            "ay_sayisi": len(monthly)
        })

    conn.close()
    return jsonify(result)

@app.route("/api/dashboard/okr-detail")
@require_auth
def dashboard_okr_detail():
    yil     = request.args.get("yil", datetime.now().year)
    company = request.args.get("company")

    if request.user["role"] == "company":
        company = request.user["company"]

    conn = get_db()
    q = "SELECT * FROM entries WHERE yil=? AND deger IS NOT NULL"
    params = [yil]
    if company:
        q += " AND company=?"; params.append(company)

    rows = [dict(r) for r in conn.execute(q, params).fetchall()]
    conn.close()

    # OKR bazında gruplama
    okr_data = {}
    for row in rows:
        key = (row["company"], row["okr_no"])
        if key not in okr_data:
            okr_data[key] = []
        okr_data[key].append(row["deger"])

    result = []
    for (comp, okr_no), vals in okr_data.items():
        okr_info = next((o for o in OKR_STRUKTUR if o["no"] == okr_no), None)
        result.append({
            "company": comp,
            "okr_no": okr_no,
            "okr_ad": okr_info["okr"] if okr_info else "",
            "avg": round(sum(vals)/len(vals), 2),
            "count": len(vals)
        })

    return jsonify(result)

# ─── API: EXCEL EXPORT ────────────────────────────────────────────────────────

@app.route("/api/export/excel")
@require_auth
def export_excel():
    yil = int(request.args.get("yil", datetime.now().year))
    company_filter = request.args.get("company")
    if request.user["role"] == "company":
        company_filter = request.user["company"]

    conn = get_db()
    q = "SELECT * FROM entries WHERE yil=?"
    params = [yil]
    if company_filter:
        q += " AND company=?"; params.append(company_filter)
    rows = [dict(r) for r in conn.execute(q, params).fetchall()]
    conn.close()

    wb = openpyxl.Workbook()

    # ── ÖZET SAYFASI ──
    ws = wb.active
    ws.title = "Özet"

    header_fill   = PatternFill("solid", fgColor="1A1E28")
    accent_fill   = PatternFill("solid", fgColor="00E5A0")
    alt_fill      = PatternFill("solid", fgColor="F0F4F8")
    header_font   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    accent_font   = Font(bold=True, color="000000", name="Calibri", size=11)
    center        = Alignment(horizontal="center", vertical="center")
    thin          = Side(style="thin", color="DDDDDD")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Başlık
    ws.merge_cells("A1:O1")
    ws["A1"] = f"KİBAR HOLDİNG – İSGÇ OKR TAKİP SİSTEMİ – {yil}"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    ws["A1"].fill = PatternFill("solid", fgColor="0A0C10")
    ws["A1"].alignment = center

    # Sütun başlıkları
    headers = ["Şirket"] + MONTH_TR + ["YTD Ort."]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = 13 if col > 1 else 22

    # Şirket verileri
    entry_map = {}
    for r in rows:
        k = (r["company"], r["ay"])
        if k not in entry_map:
            entry_map[k] = []
        if r["deger"] is not None:
            entry_map[k].append(r["deger"])

    for i, company in enumerate(COMPANIES):
        row_idx = i + 3
        ws.cell(row=row_idx, column=1, value=company).font = Font(bold=True, name="Calibri")
        ws.cell(row=row_idx, column=1).border = border

        monthly_vals = []
        for j, ay in enumerate(MONTHS):
            vals = entry_map.get((company, ay), [])
            avg = round(sum(vals)/len(vals), 2) if vals else None
            if avg is not None:
                monthly_vals.append(avg)
            cell = ws.cell(row=row_idx, column=j+2, value=avg)
            cell.alignment = center
            cell.border = border
            if i % 2 == 1:
                cell.fill = alt_fill
            if avg is not None:
                color = "00C87A" if avg >= 70 else ("FFC107" if avg >= 50 else "FF6B6B")
                cell.font = Font(color=color, bold=True, name="Calibri")

        ytd = round(sum(monthly_vals)/len(monthly_vals), 2) if monthly_vals else None
        ytd_cell = ws.cell(row=row_idx, column=14, value=ytd)
        ytd_cell.alignment = center
        ytd_cell.border = border
        ytd_cell.font = Font(bold=True, name="Calibri", color="00C87A" if (ytd or 0)>=70 else "FF6B6B")

    ws.freeze_panes = "B3"

    # ── DETAY SAYFALARI ──
    companies_to_export = [company_filter] if company_filter else COMPANIES
    for company in companies_to_export:
        safe_name = company[:30].replace("/","").replace("\\","")
        wd = wb.create_sheet(safe_name)

        wd.merge_cells("A1:G1")
        wd["A1"] = f"{company} – {yil} İSGÇ OKR Detayı"
        wd["A1"].font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        wd["A1"].fill = PatternFill("solid", fgColor="12151C")
        wd["A1"].alignment = center

        det_headers = ["OKR No", "OKR Kategorisi", "KR No", "Key Result (KR)", "Birim"] + MONTH_TR[:6] + MONTH_TR[6:]
        for col, h in enumerate(det_headers, 1):
            cell = wd.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            wd.column_dimensions[get_column_letter(col)].width = 15 if col > 2 else (8 if col==1 else 35)

        row_idx = 3
        for okr in OKR_STRUKTUR:
            for kr in okr["krs"]:
                cells_data = [okr["no"], okr["okr"], kr["no"], kr["kr"], kr["birim"]]
                for col, val in enumerate(cells_data, 1):
                    cell = wd.cell(row=row_idx, column=col, value=val)
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
                    if col <= 2:
                        cell.fill = PatternFill("solid", fgColor="E8F4FD")

                for j, ay in enumerate(MONTHS):
                    entry = next((r for r in rows if r["company"]==company and r["ay"]==ay
                                  and r["okr_no"]==okr["no"] and r["kr_no"]==kr["no"]), None)
                    val = entry["deger"] if entry else None
                    cell = wd.cell(row=row_idx, column=6+j, value=val)
                    cell.alignment = center
                    cell.border = border
                    if val is not None:
                        color = "00C87A" if val >= 70 else ("FFC107" if val >= 50 else "FF6B6B")
                        cell.font = Font(color=color, bold=True, name="Calibri")

                wd.row_dimensions[row_idx].height = 35
                row_idx += 1

        wd.freeze_panes = "F3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"ISGC_OKR_{yil}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─── YÖNETICI: KULLANICI YÖNETİMİ ────────────────────────────────────────────

@app.route("/api/admin/users", methods=["GET"])
@require_admin
def list_users():
    conn = get_db()
    users = [dict(r) for r in conn.execute(
        "SELECT id,username,role,company,created_at FROM users ORDER BY id"
    ).fetchall()]
    conn.close()
    return jsonify(users)

@app.route("/api/admin/users", methods=["POST"])
@require_admin
def create_user():
    data = request.json or {}
    username = data.get("username","").strip()
    password = data.get("password","")
    role     = data.get("role","company")
    company  = data.get("company")

    if not username or not password:
        return jsonify({"error": "Kullanıcı adı ve şifre zorunlu"}), 400

    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO users(username,password_hash,role,company) VALUES(?,?,?,?)",
            (username, hash_password(password), role, company)
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"error": "Bu kullanıcı adı zaten var"}), 409
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/admin/users/<int:uid>", methods=["DELETE"])
@require_admin
def delete_user(uid):
    conn = get_db()
    conn.execute("DELETE FROM users WHERE id=?", (uid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/admin/users/<int:uid>/password", methods=["PUT"])
@require_admin
def reset_password(uid):
    data = request.json or {}
    new_pw = data.get("password","")
    if len(new_pw) < 6:
        return jsonify({"error": "Şifre en az 6 karakter olmalı"}), 400
    conn = get_db()
    conn.execute("UPDATE users SET password_hash=? WHERE id=?", (hash_password(new_pw), uid))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

# ─── BAŞLAT ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    init_db()
    print("\n" + "="*60)
    print("  Kibar Holding – İSGÇ OKR Takip Sistemi")
    print("  http://localhost:5000")
    print("="*60)
    print("\n  Varsayılan kullanıcılar:")
    print("  admin       / Admin2026!   (yönetici)")
    print("  assan_al    / AlPass26!    (Assan Alüminyum)")
    print("  assan_hanil / HanPass26!   (Assan Hanil)")
    print("  assan_liman / LimPass26!   (Assan Liman)")
    print("  assan_loj   / LojPass26!   (Assan Lojistik)")
    print("  assan_panel / PanPass26!   (Assan Panel)")
    print("  ispak       / IspPass26!   (İspak Esnek Ambalaj)")
    print("="*60 + "\n")
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
